# -*- coding: utf-8 -*-
# @Time    : 2024-7-10 15:58
# @Author  : liangxiaopeng
# @File    : main.py
import openpyxl

# 打开源Excel文件
source_workbook = openpyxl.load_workbook('E:/py_output/edit_excel/main1.xlsx')
source_sheet = source_workbook.active

# 创建目标Excel文件
target_workbook = openpyxl.Workbook()
target_sheet = target_workbook.active

# 选择工作表
sheet = source_workbook.active

# 读取指定单元格的数据+title
cell_value11 = sheet.cell(row=1, column=1).value
cell_value13 = sheet.cell(row=1, column=3).value
cell_value21 = sheet.cell(row=2, column=1).value
cell_value23 = sheet.cell(row=2, column=3).value
cell_value31 = sheet.cell(row=3, column=1).value
cell_value33 = sheet.cell(row=3, column=3).value

print(f'单元格(cell_value11)的值为： {cell_value11}')
print(f'单元格(cell_value12)的值为： {cell_value13}')
print(f'单元格(cell_value21)的值为： {cell_value21}')
print(f'单元格(cell_value23)的值为： {cell_value23}')
print(f'单元格(cell_value31)的值为： {cell_value31}')
print(f'单元格(cell_value33)的值为： {cell_value33}')

# 读取指定单元格的数据+内容
cell_value12 = sheet.cell(row=1, column=2).value
cell_value14 = sheet.cell(row=1, column=4).value
cell_value22 = sheet.cell(row=2, column=2).value
cell_value24 = sheet.cell(row=2, column=4).value
cell_value32 = sheet.cell(row=3, column=2).value
cell_value34 = sheet.cell(row=3, column=4).value

# 写入新excel数据+tiele
target_sheet.cell(row=1, column=1).value = cell_value11
target_sheet.cell(row=1, column=2).value = cell_value13
target_sheet.cell(row=1, column=3).value = cell_value21
target_sheet.cell(row=1, column=4).value = cell_value23
target_sheet.cell(row=1, column=5).value = cell_value31
target_sheet.cell(row=1, column=6).value = cell_value33

# 写入新excel数据+内容
target_sheet.cell(row=2, column=1).value = cell_value12
target_sheet.cell(row=2, column=2).value = cell_value14
target_sheet.cell(row=2, column=3).value = cell_value22
target_sheet.cell(row=2, column=4).value = cell_value24
target_sheet.cell(row=2, column=5).value = cell_value32
target_sheet.cell(row=2, column=6).value = cell_value34

# 保存到新的excel文件中
target_workbook.save("E:/py_output/edit_excel/main2.xlsx")

# 关闭工作簿
source_workbook.close()
target_workbook.close()