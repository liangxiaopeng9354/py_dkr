# -*- coding: utf-8 -*-
# @Time    : 2024-7-10 15:58
# @Author  : liangxiaopeng
# @File    : main.py
import openpyxl

# 打开源Excel文件
source_workbook = openpyxl.load_workbook('E:/py_output/edit_excel/main1.xlsx')
source_sheet = source_workbook.active

# 创建目标Excel文件
target_workbook = openpyxl.Workbook()
target_sheet = target_workbook.active

# 读取指定单元格的数据+title
cell_values = [
    (1, 1), (1, 3), (2, 1), (2, 3), (3, 1), (3, 3)
]

for row, col in cell_values:
    cell_value = source_sheet.cell(row=row, column=col).value
    print(f'单元格({row}, {col})的值为： {cell_value}')
    target_sheet.cell(row=row, column=col).value = cell_value

# 读取指定单元格的数据+内容
cell_values = [
    (1, 2), (1, 4), (2, 2), (2, 4), (3, 2), (3, 4)
]

for row, col in cell_values:
    cell_value = source_sheet.cell(row=row, column=col).value
    target_sheet.cell(row=row, column=col + 1).value = cell_value

# 保存到新的excel文件中
target_workbook.save("E:/py_output/edit_excel/main3.xlsx")

# 关闭工作簿
source_workbook.close()
target_workbook.close()
